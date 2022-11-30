import openpyxl as op
import xlsxwriter as xl
import pandas as pd
import matplotlib.pyplot as plt
import re
import random

from openpyxl.chart import BarChart, BarChart3D , Series, Reference , PieChart
from openpyxl.worksheet.table import Table, TableStyleInfo

if __name__ == '__main__':
    
    # https://openpyxl.readthedocs.io/en/stable/worksheet_tables.html
    
    book = op.Workbook()
    
    sheet_ejec = book.active
    sheet_ejec.title = 'reporte_ejecutivo'
    sheet_ing = book.create_sheet('reporte_ingredientes')
    sheet_ped = book.create_sheet('reporte_pedidos')
    
    # CARGAR DATOS 
    
    df_pizzas = pd.read_csv('pizzas.csv', encoding='latin')
    df_tipos = pd.read_csv('pizza_types.csv', encoding='latin')
    df_orders = pd.read_csv('orders_clean.csv', encoding='latin')
    df_odetails = pd.read_csv('order_details_clean.csv', encoding='latin')
    df_analisis_tmp = pd.read_csv('analisis_pedidos_semanales.csv', encoding='latin')
    df_rec_tmp = pd.read_csv('recomendacion_ingredientes.csv', encoding='latin')
    
    # REPORTE EJECUTIVO

    # Creo el gráfico de sectores de proporciones de pizzas de distintos tamaños
    
    total_p = 0
    names = ['Pizzas_S','Pizzas_M','Pizzas_L','Pizzas_XL','Pizzas_XXL']
    
    sheet_ejec[f'I2'] = 'Tamaño'
    sheet_ejec[f'J2'] = 'Cantidad'
    
    for index in range(3,8):
        
        name = re.sub('_',' ',names[index-3])
        suma = (df_analisis_tmp[names[index-3]]).sum()
        
        sheet_ejec[f'I{index}'] = name
        sheet_ejec[f'J{index}'] = suma

        total_p += suma
    
    tabla_tam = Table(displayName="Table0", ref=f"I2:J7")

    estilo= TableStyleInfo(name="TableStyleMedium6", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabla_tam.tableStyleInfo = estilo

    sheet_ejec.add_table(tabla_tam)
        
    g_tam = PieChart()
    tams = Reference(sheet_ejec, min_col=9, min_row=3, max_row=7)
    datos = Reference(sheet_ejec, min_col=10, min_row=3, max_row=7)
    
    g_tam.add_data(datos, titles_from_data=True)
    g_tam.set_categories(tams)
    g_tam.title = "Pizzas vendidas por tamaño"
    
    sheet_ejec.add_chart(g_tam, "L2")


    # Utilizamos los datos para calcular el dinero generado por cada tipo concreto de pizza

    precios = {}
    tipos_p = list(df_pizzas['pizza_type_id'])
    sizes_p = list(df_pizzas['size'])
    prices_p = list(df_pizzas['price'])
    
    for tipo in tipos_p:
        precios[tipo] = [0,0,0,0,0]
    
    for index in range(0,len(tipos_p)):
        tipo = tipos_p[index]
        size = sizes_p[index]
        price = prices_p[index]
        
        if size == 'S':
            precios[tipo][0] = price
        if size == 'M':
            precios[tipo][1] = price
        if size == 'L':
            precios[tipo][2] = price
        if size == 'XL':
            precios[tipo][3] = price
        if size == 'XXL':
            precios[tipo][4] = price
        
    print(precios)
         
    sheet_ejec['B2'] = 'Pizza'
    sheet_ejec['C2'] = 'Precio de S'
    sheet_ejec['D2'] = 'Precio de M'
    sheet_ejec['E2'] = 'Precio de L'
    sheet_ejec['F2'] = 'Precio de XL'
    sheet_ejec['G2'] = 'Precio de XXL'
    
    count = 3
    for key in precios.keys():
        
        sheet_ejec[f'B{count}'] = df_tipos['name'][count-3]
        prec = precios.get(key)
        
        sheet_ejec[f'C{count}'] = prec[0]
        sheet_ejec[f'D{count}'] = prec[1]
        sheet_ejec[f'E{count}'] = prec[2]
        sheet_ejec[f'F{count}'] = prec[3]
        sheet_ejec[f'G{count}'] = prec[4]
        count += 1
        
    tabla_p = Table(displayName="Tablep", ref=f'B2:G{len(df_tipos["name"])+2}')

    estilo= TableStyleInfo(name="TableStyleMedium6", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabla_p.tableStyleInfo = estilo

    sheet_ejec.add_table(tabla_p)
    
    
    # Calculamos el dinero generado por tipo de pizza
    
    sheet_ejec['B36'] = 'Pizza'
    sheet_ejec['C36'] = 'Cantidad de S'
    sheet_ejec['D36'] = 'Cantidad de M'
    sheet_ejec['E36'] = 'Cantidad de L'
    sheet_ejec['F36'] = 'Cantidad de XL'
    sheet_ejec['G36'] = 'Cantidad de XXL'
    sheet_ejec['H36'] = 'Dinero generado'
    
    
    for index in range(37, len(df_analisis_tmp['Tipo_pizza'])+37):

        pizza_tmp = (df_analisis_tmp['Tipo_pizza'])[index-37]
          
        pizza_name = (df_tipos['name'])[index-37]

        sheet_ejec[f'B{index}'] = pizza_name
        sheet_ejec[f'C{index}'] = df_analisis_tmp['Pizzas_S'][index-37]
        sheet_ejec[f'D{index}'] = df_analisis_tmp['Pizzas_M'][index-37]
        sheet_ejec[f'E{index}'] = df_analisis_tmp['Pizzas_L'][index-37]
        sheet_ejec[f'F{index}'] = df_analisis_tmp['Pizzas_XL'][index-37]
        sheet_ejec[f'G{index}'] = df_analisis_tmp['Pizzas_XXL'][index-37]
        
        total = 0
        for subindex in range(0,len(names)):
            name = names[subindex]
            total += (df_analisis_tmp[name][index-37])*((precios[pizza_tmp])[subindex])
            
        sheet_ejec[f'H{index}'] = total
    
    
    tabla_price = Table(displayName="Table", ref=f"B36:H{len(df_analisis_tmp['Tipo_pizza']) +36}")

    estilo= TableStyleInfo(name="TableStyleMedium6", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabla_price.tableStyleInfo = estilo

    sheet_ejec.add_table(tabla_price)
    
    # Creo el gráfico de barras del dinero que generan las pizzas
    
    g_money = BarChart()
    g_money.type = "bar"
    g_money.style = 7
    g_money.title = "Dinero generado por tipo de pizza"
    g_money.y_axis.title = 'Pizza'
    g_money.x_axis.title = 'Beneficio'
    
    values = Reference(sheet_ejec, min_col=8, min_row=36, max_row=len(df_analisis_tmp['Tipo_pizza'])+36, max_col=8)
    pizza_names = Reference(sheet_ejec, min_col=2, max_col = 2, min_row=36, max_row=len(df_analisis_tmp['Tipo_pizza'])+36)
    
    g_money.add_data(values, titles_from_data=False)
    g_money.set_categories(pizza_names)
    g_money.legend = None
    g_money.size = (30,50)
    
    sheet_ejec.add_chart(g_money, "L18")
  
    

    # REPORTE INGREDIENTES
    
    df_reco = df_rec_tmp.sort_values('Unidades a comprar',ascending=False)
    
    sheet_ing['B2'] = 'Ingredientes'
    sheet_ing['C2'] = ' Recomendación de cantidad de compra'
    
    ing = list(df_reco['Ingredientes'])
    values = list(df_reco['Unidades a comprar'])

    for index in range(0,len(ing)):
        
        sheet_ing[f'B{index +3}'] = ing[index]
        sheet_ing[f'C{index +3}'] = int(values[index])

    # Creamos la tabla de ingredientes recomendados
    
    tabla_rec = Table(displayName="Table1", ref=f"B2:C{len(ing)+2}")

    estilo= TableStyleInfo(name="TableStyleMedium6", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabla_rec.tableStyleInfo = estilo

    sheet_ing.add_table(tabla_rec)
    
    # Gráfico de ingredientes y cantidad recomendada
    
    g_ing = BarChart()
    g_ing.type = "bar"
    g_ing.style = 7
    g_ing.title = "Recomendacion de ingredientes por semana"
    g_ing.y_axis.title = 'Ingredientes'
    g_ing.x_axis.title = 'Cantidad'
    
    data = Reference(sheet_ing, min_col=3, min_row=2, max_row=len(df_reco['Ingredientes'])+2, max_col=3)
    ingredientes = Reference(sheet_ing, min_col=2, max_col = 2, min_row=3, max_row=len(df_reco['Ingredientes'])+2)
    
    g_ing.add_data(data, titles_from_data=False)
    g_ing.set_categories(ingredientes)
    g_ing.legend = None
    g_ing.size = (30,50)
    
    sheet_ing.add_chart(g_ing, "F2")
    
    # Creo un grafico para los 5 ingredientes más utilizados
    
    g_top5 = BarChart3D()
    g_top5.type = "col"
    g_top5.style = 7
    g_top5.title = "Top 5 ingredientes más utilizados"
    g_top5.y_axis.title = 'Ingredientes'
    g_top5.x_axis.title = 'Cantidad'
    
    data_top = Reference(sheet_ing, min_col=3, min_row=3, max_row=7, max_col=3)
    ingredientes_top = Reference(sheet_ing, min_col=2, max_col = 2, min_row=3, max_row=7)
    
    g_top5.add_data(data_top, titles_from_data=False)
    g_top5.set_categories(ingredientes_top)
    g_top5.legend = None
    g_top5.size = (30,50)
    
    sheet_ing.add_chart(g_top5, "F19")
    

    # Creo un grafico para los 5 ingredientes menos utilizados
    
    g_low5 = BarChart3D()
    g_low5.type = "col"
    g_low5.style = 7
    g_low5.title = "Top 5 ingredientes menos utilizados"
    g_low5.y_axis.title = 'Ingredientes'
    g_low5.x_axis.title = 'Cantidad'
    
    data_low = Reference(sheet_ing, min_col=3, min_row=len(ing)-3, max_row=len(ing)+2, max_col=3)
    ingredientes_low = Reference(sheet_ing, min_col=2, max_col = 2, min_row=len(ing)-3, max_row=len(ing)+2)
    
    g_low5.add_data(data_low, titles_from_data=False)
    g_low5.set_categories(ingredientes_low)
    g_low5.legend = None
    g_low5.size = (30,50)
    
    sheet_ing.add_chart(g_low5, "F36")
    
    
    # REPORTE PEDIDOS
    
    df_modas = df_analisis_tmp.sort_values('Moda_anual',ascending=False)
    
    # Cambio los nombre de las pizzas para cambiar el id dado por su nombre normal
    names = {}
    for index in range(len(df_tipos['pizza_type_id'])):
        id = df_tipos['pizza_type_id'][index]
        name = df_tipos['name'][index]
        name = re.sub('The ','', name)
        name = re.sub(' Pizza','', name)
        names[id] = name
    
    for index in range(len(df_modas['Tipo_pizza'])):
        id = df_modas['Tipo_pizza'][index]
        df_modas['Tipo_pizza'][index] = names[id]
    
    sheet_ped['B2'] = 'Tipos de pizza'
    sheet_ped['C2'] = 'Moda anual'
    sheet_ped['D2'] = 'Pedidos anuales'
    sheet_ped['E2'] = 'Porcentajes por pizza'
    
    tipos = list(df_modas['Tipo_pizza'])
    modas = list(df_modas['Moda_anual'])
    pedidos = list(df_modas['Pedidos_anuales'])
    porcentajes = list(df_modas['Porcentajes_anuales(%)'])
    
    for index in range(0,len(tipos)):
        sheet_ped[f'B{index +3}'] = tipos[index]
        sheet_ped[f'C{index +3}'] = modas[index]
        sheet_ped[f'D{index +3}'] = pedidos[index]
        sheet_ped[f'E{index +3}'] = porcentajes[index]
        
    # Tabla que incluye la moda anual, el número de pedidos anuales y el porcentaje dentro de esta ultima medida por pizzas
    
    tabla_mod = Table(displayName="Table2", ref=f"B2:E{len(tipos)+2}")

    estilo= TableStyleInfo(name="TableStyleMedium6", showFirstColumn=True, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabla_mod.tableStyleInfo = estilo

    sheet_ped.add_table(tabla_mod)
    
    # Gráfica modas anuales por pizza
    
    g_mod = BarChart()
    g_mod.type = "bar"
    g_mod.style = 7
    g_mod.title = "Modas anuales por pizzas"
    g_mod.y_axis.title = 'Pizza'
    g_mod.x_axis.title = 'Moda'
    
    data_m = Reference(sheet_ped, min_col=3, min_row=2, max_row=len(tipos)+2, max_col=3)
    ingredientes_m = Reference(sheet_ped, min_col=2, max_col = 2, min_row=3, max_row=len(tipos)+2)
    
    g_mod.add_data(data_m, titles_from_data=False)
    g_mod.set_categories(ingredientes_m)
    g_mod.legend = None
    g_mod.size = (30,50)
    
    sheet_ped.add_chart(g_mod, "G2")
    
    # Gráfica de número de pedidos anuales
    
    g_anual = BarChart()
    g_anual.type = 'col'
    g_anual.style = 7
    g_anual.title = "Pedidos anuales por pizzas"
    g_anual.y_axis.title = 'Pedidos'
    
    data_p = Reference(sheet_ped, min_col=4, min_row=2, max_row=len(tipos)+2, max_col=4)
    
    g_anual.add_data(data_p, titles_from_data=False)
    g_anual.set_categories(ingredientes_m)
    g_anual.legend = None
    g_anual.size = (30,50)
    
    sheet_ped.add_chart(g_anual, "G20")
    
    
    # Guardamos el fichero
    book.save('MP_Excel_report.xlsx')